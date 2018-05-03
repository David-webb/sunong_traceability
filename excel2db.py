#!/usr/bin/python
# -*- coding: utf-8 -*-

# Created by David Teng on 18-4-24
"""
    该程序用于从苏农的化肥excel表格中提取数据，使用的是openpyxl模块
    openpyxl的读写目标是office 2010 excel，也就是.xlsx文件
    安装: pip install openpyxl
"""
import sys
import json
import pymongo
import traceback
from openpyxl import load_workbook

MONGODB_CONFIG = {
    'host': '127.0.0.1',
    'port': 27017,
    'db_name': 'envdata_db',
    'username': None,
    'password': None
}
class mongops():
    """
        提供mongo数据库的操作
    """
    def __init__(self):
        # connect db
        try:
            self.conn = pymongo.MongoClient(MONGODB_CONFIG['host'], MONGODB_CONFIG['port'])
            self.db = self.conn[MONGODB_CONFIG['db_name']]  # connect db
            self.username = MONGODB_CONFIG['username']
            self.password = MONGODB_CONFIG['password']
            if self.username and self.password:
                self.connected = self.db.authenticate(self.username, self.password)
            else:
                self.connected = True
        except Exception:
            print traceback.format_exc()
            print 'Connect Statics Database Fail.'
            sys.exit(1)

    def getsupplierId(self, name):
        """

        :return:
        """
        sid = self.db["suppliers_collection"].find_one({"name":name},{"_id":1})
        if sid:
            return sid
        else:
            return self.db["orders_collection"].find_one({"name":name}, {"supplier_id":1})
        pass

    def insertintoorders(self, itemlist, mode='m'):
        if mode == 'm':
            self.db['orders_collection'].insert_many(itemlist)
        else:
            self.db['orders_collection'].insert_one(itemlist)
        pass

    def getorderscount(self):
        return self.db['orders_collection'].find({}).count()

class getdatafromexcel_sunong():
    """
        从苏农化肥进销存excel表格中取出数据
    """
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(filename=self.filename)

    def getsheetobjlist(self):
        """
           获取excel文件中所有的sheet对象
        :return:
        """
        wb = load_workbook(filename=self.filename)
        return wb.sheetnames
        pass


    def getsheetdata(self):
        """
            获取所有sheet的数据，以字典返回
            格式: dic = {"sheetname":[[],[],[]...]}
        :return:
        """
        sheetlist = self.wb.sheetnames
        sheetsdic = {}
        for sheet in sheetlist:
            rowsdatalist = []
            sheetobj = self.wb[sheet]
            # print sheet, len(list(sheetobj.rows)[1:-1]), len(list(sheetobj.columns)), sheetobj['A1'].value
            for row in list(sheetobj.rows)[1:-1]:
                rowsdatalist.append([cell.value for cell in row])
            # shname = unicode(sheet, "utf-8")
            sheetsdic[sheet] = rowsdatalist
        return sheetsdic
        # self.savedata()
        pass

    def savedata(self, datadic, savepath):
        """
            将提取出来的excel文件数据以json保存到文件

        :return:
        """
        with open(savepath, "w")as wr:
            wr.write(json.dumps(datadic))
    pass



    def extractproductnames(self):
        """
            将粗略统计的太仓/常州的每日每种产品的进出数据
        """
        location = self.filename.split('/')[-1]
        sheetlist = self.wb.sheetnames
        productsdic = {}
        for sheet in sheetlist:
            sheetobj = self.wb[sheet]
            rowsNum = len(list(sheetobj.rows))
            for i in range(1, rowsNum):
                tmproname = sheetobj[('A%s' % i)].value
                if tmproname not in productsdic.keys():
                    productsdic[tmproname] = [sheet, sheetobj[('C%s' % i)].value, location]

        # 输出产品字典{"ProName":[首次出现的日期，期初总量],....}
        # print len(productsdic.keys())
        # for k, v in productsdic.items():
        #     print k, v[0], type(v[1]),v[1], v[2]

        # 与已有的产品类别进行对比
        # tmpdb = mongops()
        # itemlist = tmpdb.db["commodity_collection"].find({}, {'name': 1})
        # for item in itemlist:
        #     if item['name'] not in productsdic.keys():
        #         print item['name']

        return productsdic


    def insertintoorders(self):
        """

        :return:
        """
        dbobj = mongops()
        companylist = ["太仓分公司总仓", "常州分公司总仓"]
        productsdic = self.extractproductnames()
        tmpordersitem = {
            "_id": "",
            "name":"",
            "supplier_id": "",
            "type": "",
            "quantity":"",
            "total_amount":"",
            "ordered_by":"",
            "ordered_at":"2017-12-31"
        }
        itemlist = []
        sId = dbobj.getorderscount() + 1
        for k, v in productsdic.items():
            tmpordersitem["_id"] = sId
            tmpordersitem['name'] = k
            tmpordersitem['supplier_id'] = dbobj.getsupplierId(k)
            # print tmpordersitem
            tmpordersitem["quantity"] = v[1]
            if "太仓" in v[2]:
                tmpordersitem["ordered_by"] = 1003
            else:
                tmpordersitem["ordered_by"] = 1007
            itemlist.append(tmpordersitem)
            print tmpordersitem
            dbobj.insertintoorders(tmpordersitem, 's')
            sId += 1
        # dbobj.insertintoorders(itemlist)

if __name__ == '__main__':
    # wb = load_workbook(filename="数据/2018年1-4月太仓.xlsx")
    # print wb.sheetnames

    # eclfile = "./数据/2018年1-4月太仓.xlsx"
    # eclfile_1 = "./数据/2018年1-4月常州.xlsx"
    # tmpobj_1 = getdatafromexcel_sunong(eclfile_1)
    # tmpobj.getsheetdata()

    # 提取excel表格数据，并以json格式保存
    # tmpdic = tmpobj.getsheetdata()
    # savepath = eclfile.replace(".xlsx", ".json")
    # tmpobj.savedata(tmpdic,savepath)

    # 提取所有的商品数据
    # eclfile_1 = "./数据/2018年1-4月常州.xlsx"
    # tmpobj_1 = getdatafromexcel_sunong(eclfile_1)
    # tmpobj_1.extractproductnames()

    exlfile_2 = "./数据/2018年1-4月常州.xlsx"
    # exlfile_2 = "./数据/2018年1-4月太仓.xlsx"
    tmpobj_2 = getdatafromexcel_sunong(exlfile_2)
    # tmpobj_2.extractproductnames()
    tmpobj_2.insertintoorders()

    pass
